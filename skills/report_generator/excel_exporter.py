
import os
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Add parent directory to path to import lib
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

from lib.db_operations import resolve_client, run_query

def create_header_style():
    """Create standard header style"""
    return {
        'font': Font(bold=True, color="FFFFFF"),
        'fill': PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid"),
        'alignment': Alignment(horizontal="center", vertical="center"),
        'border': Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    }

def set_column_widths(ws, widths):
    """Set column widths"""
    for col_char, width in widths.items():
        ws.column_dimensions[col_char].width = width

from neo4j.time import Date as Neo4jDate

def sanitize_value(value):
    """Convert Neo4j types to Excel-compatible types"""
    if isinstance(value, Neo4jDate):
        return value.iso_format()
    return value

def export_client_data_to_excel(client_identifier: str) -> str:
    """
    Export client data to an Excel file.
    
    Args:
        client_identifier: Name or ID of the client.
        
    Returns:
        Absolute path to the generated Excel file.
    """
    client = resolve_client(client_identifier)
    if not client:
        raise ValueError(f"Client not found: {client_identifier}")
        
    client_name = client.get('name') or client.get('displayCode')
    
    # Create Workbook
    wb = Workbook()
    
    # --- Sheet 1: Profile ---
    ws_profile = wb.active
    ws_profile.title = "Profile"
    
    # Title
    ws_profile['A1'] = f"Client Profile: {client_name}"
    ws_profile['A1'].font = Font(size=16, bold=True)
    
    # Basic Info
    ws_profile['A3'] = "Basic Information"
    ws_profile['A3'].font = Font(bold=True)
    
    basic_info = [
        ("Name", client_name),
        ("DOB", client.get('dob', 'Unknown')),
        ("Blood Type", client.get('bloodType', 'Unknown')),
        ("Display Code", client.get('displayCode', '-'))
    ]
    
    for i, (key, value) in enumerate(basic_info, start=4):
        ws_profile[f'A{i}'] = key
        ws_profile[f'B{i}'] = sanitize_value(value)
        
    # Certificates
    ws_profile['A9'] = "Certificates"
    ws_profile['A9'].font = Font(bold=True)
    
    header_style = create_header_style()
    headers = ["Type", "Grade", "Next Renewal"]
    for col, header in enumerate(headers, start=1):
        cell = ws_profile.cell(row=10, column=col, value=header)
        for k, v in header_style.items():
            setattr(cell, k, v)
            
    certs = run_query("""
        MATCH (c:Client)-[:HAS_CERTIFICATE]->(cert:Certificate)
        WHERE c.clientId = $cid OR c.name = $name
        RETURN cert.type as type, cert.grade as grade, cert.nextRenewalDate as renewal
        ORDER BY cert.nextRenewalDate
    """, {"cid": client.get('clientId'), "name": client_name})
    
    for row_idx, cert in enumerate(certs, start=11):
        ws_profile.cell(row=row_idx, column=1, value=sanitize_value(cert.get('type')))
        ws_profile.cell(row=row_idx, column=2, value=sanitize_value(cert.get('grade')))
        ws_profile.cell(row=row_idx, column=3, value=sanitize_value(cert.get('renewal')))

    # Key Persons
    start_row = 10 + len(certs) + 3
    ws_profile[f'A{start_row}'] = "Key Persons"
    ws_profile[f'A{start_row}'].font = Font(bold=True)
    
    start_row += 1
    headers = ["Name", "Relationship", "Phone", "Role"]
    for col, header in enumerate(headers, start=1):
        cell = ws_profile.cell(row=start_row, column=col, value=header)
        for k, v in header_style.items():
            setattr(cell, k, v)
            
    kps = run_query("""
        MATCH (c:Client)-[r:HAS_KEY_PERSON]->(kp:KeyPerson)
        WHERE c.clientId = $cid OR c.name = $name
        RETURN kp.name as name, kp.relationship as rel, kp.phone as phone, kp.role as role
        ORDER BY r.rank
    """, {"cid": client.get('clientId'), "name": client_name})
    
    for row_idx, kp in enumerate(kps, start=start_row+1):
        ws_profile.cell(row=row_idx, column=1, value=kp.get('name'))
        ws_profile.cell(row=row_idx, column=2, value=kp.get('rel'))
        ws_profile.cell(row=row_idx, column=3, value=kp.get('phone'))
        ws_profile.cell(row=row_idx, column=4, value=kp.get('role'))
        
    set_column_widths(ws_profile, {'A': 20, 'B': 15, 'C': 20, 'D': 20})

    # --- Sheet 2: Care Knowledge ---
    ws_care = wb.create_sheet("Care Knowledge")
    
    # NgActions (Contraindications)
    ws_care['A1'] = "Contraindications (NgActions)"
    ws_care['A1'].font = Font(size=14, bold=True, color="FF0000")
    
    headers = ["Action", "Risk Level", "Reason"]
    for col, header in enumerate(headers, start=1):
        cell = ws_care.cell(row=3, column=col, value=header)
        for k, v in header_style.items():
            setattr(cell, k, v)
            
    ng_actions = run_query("""
        MATCH (c:Client)-[:MUST_AVOID]->(ng:NgAction)
        WHERE c.clientId = $cid OR c.name = $name
        RETURN ng.action as action, ng.riskLevel as risk, ng.reason as reason
        ORDER BY ng.riskLevel
    """, {"cid": client.get('clientId'), "name": client_name})
    
    for row_idx, ng in enumerate(ng_actions, start=4):
        ws_care.cell(row=row_idx, column=1, value=ng.get('action'))
        ws_care.cell(row=row_idx, column=2, value=ng.get('risk'))
        ws_care.cell(row=row_idx, column=3, value=ng.get('reason'))
        
    # Care Preferences
    start_row = 4 + len(ng_actions) + 3
    ws_care[f'A{start_row}'] = "Care Preferences"
    ws_care[f'A{start_row}'].font = Font(size=14, bold=True)
    
    start_row += 1
    headers = ["Category", "Instruction", "Priority"]
    for col, header in enumerate(headers, start=1):
        cell = ws_care.cell(row=start_row, column=col, value=header)
        for k, v in header_style.items():
            setattr(cell, k, v)
            
    care_prefs = run_query("""
        MATCH (c:Client)-[:REQUIRES]->(cp:CarePreference)
        WHERE c.clientId = $cid OR c.name = $name
        RETURN cp.category as cat, cp.instruction as inst, cp.priority as pri
        ORDER BY cp.priority
    """, {"cid": client.get('clientId'), "name": client_name})
    
    for row_idx, cp in enumerate(care_prefs, start=start_row+1):
        ws_care.cell(row=row_idx, column=1, value=cp.get('cat'))
        ws_care.cell(row=row_idx, column=2, value=cp.get('inst'))
        ws_care.cell(row=row_idx, column=3, value=cp.get('pri'))

    set_column_widths(ws_care, {'A': 30, 'B': 20, 'C': 40})

    # --- Sheet 3: Support History ---
    ws_history = wb.create_sheet("Support History")
    ws_history['A1'] = "Recent Support Logs (Last 50)"
    ws_history['A1'].font = Font(size=14, bold=True)
    
    headers = ["Date", "Supporter", "Situation", "Action", "Effectiveness"]
    for col, header in enumerate(headers, start=1):
        cell = ws_history.cell(row=3, column=col, value=header)
        for k, v in header_style.items():
            setattr(cell, k, v)
            
    logs = run_query("""
        MATCH (s:Supporter)-[:LOGGED]->(log:SupportLog)-[:ABOUT]->(c:Client)
        WHERE c.clientId = $cid OR c.name = $name
        RETURN log.date as date, s.name as supporter, log.situation as sit, log.action as act, log.effectiveness as eff
        ORDER BY log.date DESC
        LIMIT 50
    """, {"cid": client.get('clientId'), "name": client_name})
    
    for row_idx, log in enumerate(logs, start=4):
        ws_history.cell(row=row_idx, column=1, value=str(log.get('date')))
        ws_history.cell(row=row_idx, column=2, value=log.get('supporter'))
        ws_history.cell(row=row_idx, column=3, value=log.get('sit'))
        ws_history.cell(row=row_idx, column=4, value=log.get('act'))
        ws_history.cell(row=row_idx, column=5, value=log.get('eff'))
        
    set_column_widths(ws_history, {'A': 15, 'B': 15, 'C': 30, 'D': 30, 'E': 15})
    
    # Save
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"ClientData_{client_name}_{timestamp}.xlsx"
    # Ensure output directory exists (using tmp or report dir in future, for now local)
    output_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "reports")
    os.makedirs(output_dir, exist_ok=True)
    
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    
    return filepath

if __name__ == "__main__":
    # Test run
    if len(sys.argv) > 1:
        name = sys.argv[1]
        print(f"Generating Excel for {name}...")
        try:
            path = export_client_data_to_excel(name)
            print(f"Success: {path}")
        except Exception as e:
            print(f"Error: {e}")
    else:
        print("Usage: python excel_exporter.py <ClientName>")
